import openpyxl as xl


# Register user details
class Database:
    def __init__(self, account_no, first_name, last_name, email, password, customerId ):
        self.account_no = account_no
        self.first_name = first_name
        self.last_name = last_name
        self.email = email
        self.password = password
        self.customerId = customerId

    # def save_to_excel(self):
    #     wb = xl.load_workbook('bank_database_excel.xlsx')
    #     sheet = wb['Sheet1']
    #
    #     self.customerId += 1
    #
    #     sheet.cell(self.customerId, 1).value = self.account_no
    #     sheet.cell(self.customerId, 2).value = self.first_name
    #     sheet.cell(self.customerId, 3).value = self.last_name
    #     sheet.cell(self.customerId, 4).value = self.email
    #     sheet.cell(self.customerId, 5).value = self.password
    #
    #     wb.save('bank_database_excel.xlsx')

    def save_to_excel(self):
        wb = xl.load_workbook('bank_database_excel.xlsx')
        sheet = wb['Sheet1']

        while sheet.cell(self.customerId, 1).value:
            self.customerId += 1
        else:
            sheet.cell(self.customerId, 1).value = self.account_no
            sheet.cell(self.customerId, 2).value = self.first_name
            sheet.cell(self.customerId, 3).value = self.last_name
            sheet.cell(self.customerId, 4).value = self.email
            sheet.cell(self.customerId, 5).value = self.password
            sheet.cell(self.customerId, 6).value = 0  # account Balance

            wb.save('bank_database_excel.xlsx')


# Authenticate user login details
class AuthDatabase:
    def __init__(self, accountNoFromUser, pswdFromUser):
        self.accountNoFromUser = accountNoFromUser
        self.pswdFromUser = pswdFromUser

    def authFromDatabase(self):
        wb = xl.load_workbook('bank_database_excel.xlsx')
        sheet = wb['Sheet1']

        convertAccountNoToList = []
        convertPasswordToList = []
        for row in range(1, sheet.max_row + 1):
            convertAccountNoToList.append(sheet.cell(row, 1).value)
            convertPasswordToList.append(sheet.cell(row, 5).value)

        if self.accountNoFromUser in convertAccountNoToList and self.pswdFromUser in convertPasswordToList:
            return True

        else:
            return False

    # To return the username
    def get_user_name(self):
        wb = xl.load_workbook('bank_database_excel.xlsx')
        sheet = wb['Sheet1']

        for account_no in range(1, sheet.max_row + 1):
            while self.accountNoFromUser != sheet.cell(account_no, 1).value:
                break
            else:
                return sheet.cell(account_no, 2).value

    def get_user_account_balance(self):
        wb = xl.load_workbook('bank_database_excel.xlsx')
        sheet = wb['Sheet1']

        for account_no in range(1, sheet.max_row + 1):
            while self.accountNoFromUser != sheet.cell(account_no, 1).value:
                break
            else:
                return sheet.cell(account_no, 6).value


class UpdateAccountBalance(AuthDatabase):
    def __init__(self, user_new_account_balance, accountNoFromUser, pswdFromUser):
        super().__init__(accountNoFromUser, pswdFromUser)
        self.user_new_account_balance = user_new_account_balance

    def update_account_balance(self):
        wb = xl.load_workbook('bank_database_excel.xlsx')
        sheet = wb['Sheet1']

        for account_no in range(1, sheet.max_row + 1):
            while self.accountNoFromUser != sheet.cell(account_no, 1).value:
                break
            else:
                sheet.cell(account_no, 6).value = self.user_new_account_balance

        wb.save('bank_database_excel.xlsx')

